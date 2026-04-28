#!/usr/bin/python3

import openpyxl
import os
import shutil


def get_brand_model_dict():

    refer_excel = openpyxl.load_workbook("language.xlsx")
    refer_sheet = refer_excel["语言包"]
    brand_model_dict = {}

    folder = os.getcwd()

    for item in range(2, refer_sheet.max_column + 1):
        name = (refer_sheet.cell(row=1, column=item)).value
        filePath = folder + "/" + name + ".txt"
        if os.path.exists(filePath):
            os.remove(filePath)

        suffix = 0
        for row in range(2, refer_sheet.max_row + 1):
            brand = (refer_sheet.cell(row=row, column=3)).value
            model = (refer_sheet.cell(row=row, column=item)).value

            if brand is None:
                print(" ")
            elif brand:
                if model is None:
                    print(" ")
                elif model:
                    file = open(filePath, "a")
                    bt1 = '"' + brand + '"'
                    suffix = 0
                    for item_1 in open(filePath):
                        tmp = item_1.find(bt1)
                        # print("item_1.find",tmp,item_1,bt1)
                        if tmp != -1:
                            suffix += 1
                            print(bt1, suffix)
                            brand = brand + "_" + str(suffix)
                    file.write('"' + brand + '"' + " = " + '"' + model + '"' ";\n")
                    file.close()


def findStrInFile(filePath, str):
    try:
        with open(filePath, "r", encoding="utf-8") as file:
            lines = []
            found = False
            for line in file:
                if not found:
                    lines.append(line)
                    if str in line:
                        found = True
                else:
                    break
        with open(filePath, "w", encoding="utf-8") as file:
            file.writelines(lines)

        if found:
            return True
        else:
            return False
    except Exception as e:
        return False


def find_file_path():
    folder = os.getcwd()
    parent_dir = os.path.dirname(folder)
    lproj_path = os.path.join(parent_dir, "JieliJianKang")

    targetlprojList = []
    for root, dirs, files in os.walk(lproj_path):
        for dir in dirs:
            if dir.endswith(".lproj"):
                targetlprojList.append(dir)

    textList = []
    for root, dirs, files in os.walk(folder):
        for file in files:
            if file.endswith(".txt"):
                textList.append(file)

    for text in textList:
        filePath = os.path.join(folder, text)
        with open(filePath, "r", encoding="utf-8") as file:
            firstLine = file.readline().strip()
            sourceContent = file.read()
            for targetlproj in targetlprojList:
                tgdir = os.path.join(lproj_path, targetlproj)
                tgfile = os.path.join(tgdir, "Localizable.strings")
                if findStrInFile(tgfile, firstLine):
                    with open(tgfile, "a", encoding="utf-8") as f:
                        f.write(sourceContent)
        os.remove(filePath)

    return


def copyToTarget(source_file, destination_file):

    with open(source_file, "r", encoding="utf-8") as f:
        content = f.read()

    if os.path.exists(destination_file):
        os.remove(destination_file)

    with open(destination_file, "w", encoding="utf-8") as f:
        f.write(content)

    shutil.copy2(source_file, destination_file)


# ~~~~~~~~~~~~~~~~main~~~~~~~~~~~~~~~~~~
# 读取对应关系表格：


def copy_file():
    get_brand_model_dict()
    find_file_path()
    print("done.")


copy_file()
